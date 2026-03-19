"""
MiniClaw Excel Tools
Handles Excel file operations for study plans and data management
"""

import os
from typing import List, Optional, Dict, Any
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from miniclaw.config.settings import settings
from miniclaw.utils.helpers import ensure_dir, format_datetime


def get_excel_path(filename: str) -> Path:
    excel_dir = ensure_dir(settings.EXCEL_DIR)
    if not filename.endswith(".xlsx"):
        filename = f"{filename}.xlsx"
    return excel_dir / filename


def create_excel(
    filename: str,
    columns: List[str],
    data: Optional[List[List]] = None,
) -> str:
    filepath = get_excel_path(filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    if data:
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(str(col_name))
        for row in ws.iter_rows(min_row=2, max_col=col_idx, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
    
    wb.save(filepath)
    return str(filepath)


def read_excel(filename: str) -> pd.DataFrame:
    filepath = get_excel_path(filename)
    
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")
    
    return pd.read_excel(filepath)


def update_cell(filename: str, row: int, column: str, value: Any) -> None:
    filepath = get_excel_path(filename)
    
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    col_idx = None
    for cell in ws[1]:
        if cell.value == column:
            col_idx = cell.column
            break
    
    if col_idx is None:
        raise ValueError(f"Column '{column}' not found in Excel file")
    
    ws.cell(row=row, column=col_idx, value=value)
    wb.save(filepath)


def analyze_excel(filename: str, analysis_type: str, column: Optional[str] = None) -> Dict[str, Any]:
    df = read_excel(filename)
    
    result = {"analysis_type": analysis_type, "total_rows": len(df)}
    
    if analysis_type == "summary":
        result["columns"] = list(df.columns)
        result["dtypes"] = df.dtypes.astype(str).to_dict()
        result["shape"] = list(df.shape)
        if column and column in df.columns:
            result["column_stats"] = df[column].describe().to_dict()
    
    elif analysis_type == "count":
        if column and column in df.columns:
            result["counts"] = df[column].value_counts().to_dict()
        else:
            result["row_count"] = len(df)
    
    elif analysis_type == "sum":
        if column and column in df.columns:
            numeric_values = pd.to_numeric(df[column], errors="coerce")
            result["sum"] = float(numeric_values.sum())
    
    elif analysis_type == "average":
        if column and column in df.columns:
            numeric_values = pd.to_numeric(df[column], errors="coerce")
            result["average"] = float(numeric_values.mean())
    
    elif analysis_type == "sort":
        if column and column in df.columns:
            result["sorted_data"] = df.sort_values(by=column).head(10).to_dict("records")
    
    return result


def create_study_excel(plan: Dict[str, Any], filename: str = "study_plan") -> str:
    columns = ["阶段", "任务", "计划日期", "完成状态", "备注"]
    
    data = []
    for stage in plan.get("stages", []):
        stage_name = stage.get("name", "")
        for task in stage.get("tasks", []):
            data.append([stage_name, task, "", "待完成", ""])
    
    if not data:
        data = [["", "", "", "", ""]]
    
    return create_excel(filename, columns, data)


def append_row(filename: str, row_data: List[Any]) -> None:
    filepath = get_excel_path(filename)
    
    if not filepath.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    ws.append(row_data)
    wb.save(filepath)


def get_all_excel_files() -> List[str]:
    excel_dir = Path(settings.EXCEL_DIR)
    if not excel_dir.exists():
        return []
    
    return [f.name for f in excel_dir.glob("*.xlsx")]
